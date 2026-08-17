"""
tests/test_statements_api.py
============================
API tests for the whole pipeline: upload, review, narrative, compare, WP-514.
Run:  python -m pytest tests/ -v
"""

import glob
import os

import pytest
from fastapi.testclient import TestClient

from backend.api import store
from backend.api.main import app

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")
REVIEW_XLSX = os.path.join(DATA, "review", "MeridianBank_FY2026.xlsx")
CLEAN_XLSX = os.path.join(DATA, "clean", "MeridianBank_FY2026_CLEAN.xlsx")
PEER_XLSX = os.path.join(DATA, "clean", "SterlingBank_FY2026_CLEAN.xlsx")
REVIEW_CSVS = sorted(glob.glob(os.path.join(DATA, "review", "csv", "*.csv")))

dataset_exists = pytest.mark.skipif(
    not os.path.exists(REVIEW_XLSX), reason="Run tools/make_dataset.py first.")


@pytest.fixture
def client():
    store.clear()
    with TestClient(app) as c:
        yield c
    store.clear()


def upload_xlsx(client, path):
    with open(path, "rb") as handle:
        r = client.post("/upload",
                        files=[("files", (os.path.basename(path), handle, "x"))])
    assert r.status_code == 200, r.text
    return r.json()["session_id"]


def upload_csvs(client, paths):
    files = [("files", (os.path.basename(p), open(p, "rb"), "text/csv"))
             for p in paths]
    r = client.post("/upload", files=files)
    assert r.status_code == 200, r.text
    return r.json()["session_id"]


# ── system ───────────────────────────────────────────────────────────────────

def test_health(client):
    assert client.get("/health").json()["status"] == "ok"


def test_formats_are_advertised_not_hardcoded(client):
    body = client.get("/formats").json()
    assert "meridian_xlsx" in body["readers"]
    assert ".csv" in body["extensions"]


def test_openapi_schema_generates(client):
    """Members browse /docs before writing frontend code."""
    paths = client.get("/openapi.json").json()["paths"]
    for path in ("/upload", "/review/{session_id}", "/compare",
                 "/wp514/{session_id}"):
        assert path in paths


# ── upload ───────────────────────────────────────────────────────────────────

@dataset_exists
def test_xlsx_upload_returns_metadata(client):
    with open(REVIEW_XLSX, "rb") as handle:
        body = client.post("/upload", files=[
            ("files", ("MeridianBank_FY2026.xlsx", handle, "x"))]).json()
    assert body["company_name"] == "Meridian Bank"
    assert body["years"] == ["FY2024", "FY2025", "FY2026"]
    assert body["label_issues_found"] == 3


@dataset_exists
def test_csv_set_upload(client):
    sid = upload_csvs(client, REVIEW_CSVS)
    assert client.post(f"/review/{sid}").json()["summary"]["FAIL"] == 12


@dataset_exists
def test_excel_and_csv_give_the_same_result(client):
    a = client.post(f"/review/{upload_xlsx(client, REVIEW_XLSX)}").json()
    b = client.post(f"/review/{upload_csvs(client, REVIEW_CSVS)}").json()
    assert a["summary"] == b["summary"]


def test_wrong_extension_is_rejected_readably(client):
    import io
    r = client.post("/upload", files=[
        ("files", ("notes.txt", io.BytesIO(b"hello"), "text/plain"))])
    assert r.status_code == 400
    assert ".txt" in r.json()["detail"]


def test_empty_file_is_rejected(client):
    import io
    r = client.post("/upload", files=[
        ("files", ("empty.xlsx", io.BytesIO(b""), "x"))])
    assert r.status_code == 400


def test_non_statement_xlsx_is_rejected_not_crashed(client):
    import io
    import openpyxl
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "hello"
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    r = client.post("/upload", files=[("files", ("random.xlsx", buffer, "x"))])
    assert r.status_code == 400


def test_directory_traversal_in_filename_is_neutralised(client):
    with open(REVIEW_XLSX, "rb") as handle:
        body = client.post("/upload", files=[
            ("files", ("../../../etc/passwd.xlsx", handle, "x"))]).json()
    assert "/" not in body["source_file"]


# ── review ───────────────────────────────────────────────────────────────────

@dataset_exists
def test_review_finds_every_planted_defect(client):
    body = client.post(f"/review/{upload_xlsx(client, REVIEW_XLSX)}").json()
    assert body["verdict"] == "EXCEPTIONS FOUND"
    assert body["summary"]["FAIL"] == 12
    assert body["summary"]["WARN"] == 2


@dataset_exists
def test_clean_file_is_clean_over_http(client):
    body = client.post(f"/review/{upload_xlsx(client, CLEAN_XLSX)}").json()
    assert body["verdict"] == "CLEAN"
    assert body["summary"]["FAIL"] == 0


@dataset_exists
def test_findings_carry_every_field_the_ui_needs(client):
    body = client.post(f"/review/{upload_xlsx(client, REVIEW_XLSX)}").json()
    for finding in body["findings"]:
        for key in ("check_id", "category", "statement", "period", "rule",
                    "expected", "found", "delta", "status", "root_cause",
                    "note"):
            assert key in finding


@dataset_exists
def test_review_is_cached(client):
    sid = upload_xlsx(client, REVIEW_XLSX)
    assert client.post(f"/review/{sid}").json() == \
           client.post(f"/review/{sid}").json()


def test_unknown_session_is_404_with_guidance(client):
    r = client.post("/review/nosuchsession")
    assert r.status_code == 404
    assert "expired" in r.json()["detail"].lower()


# ── narrative ────────────────────────────────────────────────────────────────

@dataset_exists
def test_narrative_works_without_an_api_key(client):
    body = client.post(f"/narrative/{upload_xlsx(client, REVIEW_XLSX)}").json()
    assert body["summary"]["sections"] == 5
    assert all(s["commentary"] for s in body["sections"])


@dataset_exists
def test_narrative_sections_carry_provenance(client):
    body = client.post(f"/narrative/{upload_xlsx(client, REVIEW_XLSX)}").json()
    for section in body["sections"]:
        assert section["source"]
        assert section["number_check"] in ("PASS", "FLAGGED")
        assert section["risk_level"] in ("LOW", "MEDIUM", "HIGH")


# ── peer comparison ──────────────────────────────────────────────────────────

@dataset_exists
def test_compare_two_companies(client):
    a = upload_xlsx(client, CLEAN_XLSX)
    b = upload_xlsx(client, PEER_XLSX)
    body = client.get("/compare", params={"a": a, "b": b}).json()

    assert len(body["companies"]) == 2
    assert body["comparison_year"] == "FY2026"
    assert body["scorecard"]["metrics_scored"] > 0


@dataset_exists
def test_profitability_and_growth_leaders_differ(client):
    """
    The comparison is only interesting if the two banks lead on different
    things. Meridian is more profitable; Sterling grows faster.
    """
    a = upload_xlsx(client, CLEAN_XLSX)
    b = upload_xlsx(client, PEER_XLSX)
    body = client.get("/compare", params={"a": a, "b": b}).json()
    leaders = {row["key"]: row["leader"] for row in body["metrics"]}
    assert leaders["roe_pct"] == "Meridian Bank"
    assert leaders["profit_cagr"] == "Sterling Bank"


@dataset_exists
def test_unscored_metrics_have_no_leader(client):
    """A higher tax rate is not a virtue."""
    a = upload_xlsx(client, CLEAN_XLSX)
    b = upload_xlsx(client, PEER_XLSX)
    body = client.get("/compare", params={"a": a, "b": b}).json()
    for row in body["metrics"]:
        if row["key"] in ("effective_tax_pct", "credit_deposit_pct"):
            assert row["leader"] is None


@dataset_exists
def test_comparing_a_session_with_itself_is_rejected(client):
    sid = upload_xlsx(client, CLEAN_XLSX)
    assert client.get("/compare", params={"a": sid, "b": sid}).status_code == 400


def test_compare_requires_both_parameters(client):
    assert client.get("/compare", params={"a": "x"}).status_code == 422


# ── sessions and workpaper ───────────────────────────────────────────────────

@dataset_exists
def test_sessions_are_listed(client):
    upload_xlsx(client, CLEAN_XLSX)
    upload_xlsx(client, PEER_XLSX)
    names = {s["company_name"] for s in client.get("/sessions").json()["sessions"]}
    assert names == {"Meridian Bank", "Sterling Bank"}


@dataset_exists
def test_session_can_be_deleted(client):
    sid = upload_xlsx(client, CLEAN_XLSX)
    assert client.delete(f"/sessions/{sid}").json()["deleted"] is True
    assert client.post(f"/review/{sid}").status_code == 404


@dataset_exists
def test_wp514_downloads_a_real_workbook(client):
    r = client.get(f"/wp514/{upload_xlsx(client, REVIEW_XLSX)}")
    assert r.status_code == 200
    assert r.content[:2] == b"PK"          # every .xlsx is a zip
    assert "WP514" in r.headers["content-disposition"]


@dataset_exists
def test_wp514_contains_every_section(client, tmp_path):
    import openpyxl
    r = client.get(f"/wp514/{upload_xlsx(client, REVIEW_XLSX)}")
    path = tmp_path / "wp.xlsx"
    path.write_bytes(r.content)
    sheets = openpyxl.load_workbook(path).sheetnames
    assert "Cover" in sheets
    assert any("Checks" in s for s in sheets)
    assert any("Narrative" in s for s in sheets)


# ── the full journey, in the order the frontend performs it ──────────────────

@dataset_exists
def test_end_to_end(client):
    a = upload_xlsx(client, REVIEW_XLSX)
    b = upload_xlsx(client, PEER_XLSX)
    assert client.post(f"/review/{a}").json()["verdict"] == "EXCEPTIONS FOUND"
    assert client.post(f"/narrative/{a}").json()["summary"]["sections"] == 5
    assert client.get("/compare", params={"a": a, "b": b}).status_code == 200
    assert client.get(f"/wp514/{a}").status_code == 200
