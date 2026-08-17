"""
backend/review/wp514.py
=======================
WP-514 generator for the fixed-schema review report. Same house style as
backend/reports/wp514.py, adapted for the A-E category structure and the
Meridian statement layout.

    from backend.review.wp514 import StatementWP514Generator
    StatementWP514Generator(review_report, narrative_report).save("out/")
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY, BLUE, LIGHT, GREY, WHITE = "1B2A4A", "2563EB", "EFF6FF", "F1F5F9", "FFFFFF"
STATUS_FILL = {"PASS": "DCFCE7", "WARN": "FEF9C3", "FAIL": "FEE2E2", "INFO": "F1F5F9"}
STATUS_TEXT = {"PASS": "166534", "WARN": "854D0E", "FAIL": "991B1B", "INFO": "475569"}
RISK_FILL = {"LOW": "DCFCE7", "MEDIUM": "FEF9C3", "HIGH": "FEE2E2"}
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _title(sheet, row, text, width=8):
    sheet.cell(row, 1, text).font = Font(bold=True, size=13, color=WHITE)
    for c in range(1, width + 1):
        sheet.cell(row, c).fill = PatternFill("solid", fgColor=NAVY)
    sheet.row_dimensions[row].height = 22
    return row + 1


def _subtitle(sheet, row, text, width=8):
    sheet.cell(row, 1, text).font = Font(bold=True, size=11, color=NAVY)
    for c in range(1, width + 1):
        sheet.cell(row, c).fill = PatternFill("solid", fgColor=LIGHT)
    return row + 1


def _header(sheet, row, headers):
    for i, text in enumerate(headers, 1):
        cell = sheet.cell(row, i, text)
        cell.font = Font(bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[row].height = 26
    return row + 1


def _widths(sheet, widths):
    for i, w in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = w


def _fmt(v):
    return "" if v is None else v


SOURCE_LABELS = {
    "template": "Engine-written",
    "template_no_api_key": "Engine-written (no AI configured)",
}


def _source_label(source: str) -> str:
    """
    Turn an internal source tag into something an auditor can read.

    The engine records exactly why it fell back — 'template_fallback
    (HTTPStatusError)' — which is the right level of detail for a log or an
    API response. In a signed workpaper it just needs to say who wrote the
    sentence and whether the AI was involved.
    """
    if source.startswith("gemini:"):
        return f"AI ({source.split(':', 1)[1]})"
    if source.startswith("template_fallback"):
        return "Engine-written (AI unavailable)"
    return SOURCE_LABELS.get(source, "Engine-written")


class StatementWP514Generator:
    def __init__(self, review: Dict[str, Any],
                 narrative: Optional[Dict[str, Any]] = None) -> None:
        self.review = review
        self.narrative = narrative
        self.prepared_at = datetime.now().strftime("%d %b %Y, %H:%M")

    def _cover(self, wb: Workbook) -> None:
        sheet = wb.active
        sheet.title = "Cover"
        _widths(sheet, [28, 32, 14, 14, 14])
        row = _title(sheet, 1, "  AUDIT WORKPAPER WP-514")
        sheet.cell(row, 1, "  Financial Statement Review — Fixed Schema Engine")
        sheet.cell(row, 1).font = Font(size=10, italic=True, color="64748B")
        row += 2

        summary = self.review["summary"]
        for label, value in [
            ("Client", self.review["company_name"]),
            ("Source file", self.review["source_file"]),
            ("Periods reviewed", ", ".join(self.review["periods_checked"])),
            ("Prepared by", "Automated review engine"),
            ("Prepared on", self.prepared_at),
            ("Rounding tolerance", f"{self.review['tolerance_crore']:g} crore"),
        ]:
            sheet.cell(row, 1, label).font = Font(bold=True, size=10)
            sheet.cell(row, 2, value).font = Font(size=10)
            row += 1

        row += 1
        row = _subtitle(sheet, row, "  REVIEW OUTCOME")
        sheet.cell(row, 1, "Verdict").font = Font(bold=True, size=10)
        v = sheet.cell(row, 2, self.review["verdict"])
        v.font = Font(bold=True, size=12,
                      color="991B1B" if summary["FAIL"] else "166534")
        row += 2

        row = _header(sheet, row, ["Result", "Count"])
        for label, count in [
            ("Total checks", summary["total_checks"]),
            ("Passed", summary["PASS"]), ("Warnings", summary["WARN"]),
            ("Failed", summary["FAIL"]), ("Informational", summary["INFO"]),
            ("Root-cause failures", summary["root_cause_failures"]),
        ]:
            sheet.cell(row, 1, label).font = Font(bold=True, size=10)
            sheet.cell(row, 2, count).font = Font(size=10)
            for c in (1, 2):
                sheet.cell(row, c).border = BORDER
            row += 1

        row += 1
        row = _subtitle(sheet, row, "  METHODOLOGY")
        for line in [
            "All arithmetic is performed deterministically in Python.",
            "Sections A-D check the numbers; Section E checks labels and narrative text.",
            "Section F narrative is AI-generated strictly from figures verified above,",
            "with every cited figure checked back against the supplied values.",
        ]:
            sheet.cell(row, 1, line).font = Font(size=9, color="475569")
            row += 1

        row += 1
        row = _subtitle(sheet, row, "  SIGN-OFF")
        for label in ("Reviewed by", "Signature", "Date"):
            sheet.cell(row, 1, label).font = Font(bold=True, size=10)
            for c in (2, 3):
                sheet.cell(row, c).border = BORDER
            sheet.row_dimensions[row].height = 24
            row += 1

    def _checks(self, wb: Workbook) -> None:
        sheet = wb.create_sheet("Checks A-E")
        _widths(sheet, [10, 22, 44, 13, 13, 12, 15, 40])
        row = _title(sheet, 1, "  SECTIONS A-E  |  MATH, CONSISTENCY, PRIOR YEAR, "
                               "RATIOS, TEXT")
        row += 1

        categories = sorted({f["category"] for f in self.review["findings"]})
        for category in categories:
            row = _subtitle(sheet, row, f"  {category}")
            row = _header(sheet, row, ["Period", "Check", "Rule", "Expected",
                                       "Found", "Difference", "Status", "Note"])
            header_row = row - 1

            for f in [x for x in self.review["findings"] if x["category"] == category]:
                status = f["status"]
                values = [f["period"], f["check_id"],
                         f["rule"] + ("" if f["root_cause"] else "  (knock-on)"),
                         _fmt(f["expected"]), _fmt(f["found"]), _fmt(f["delta"]),
                         status, f["note"]]
                for i, v in enumerate(values, 1):
                    cell = sheet.cell(row, i, v)
                    cell.border = BORDER
                    cell.font = Font(size=9)
                    if i in (4, 5, 6):
                        cell.number_format = "#,##0.00"
                    if i == 7:
                        cell.fill = PatternFill("solid", fgColor=STATUS_FILL.get(status, GREY))
                        cell.font = Font(size=9, bold=True, color=STATUS_TEXT.get(status, "000"))
                        cell.alignment = Alignment(horizontal="center")
                    if i == 8:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        cell.font = Font(size=8, color="64748B")
                row += 1
            sheet.auto_filter.ref = f"A{header_row}:H{row - 1}"
            row += 1
        sheet.freeze_panes = "A3"

    def _narrative_sheet(self, wb: Workbook) -> None:
        if not self.narrative:
            return
        sheet = wb.create_sheet("Narrative F")
        _widths(sheet, [8, 30, 90, 12, 24, 14])
        row = _title(sheet, 1, "  SECTION F  |  REVIEW COMMENTARY", 6)
        row += 1
        summary = self.narrative["summary"]
        state = (f"Model: {self.narrative['model']}" if self.narrative["ai_enabled"]
                 else "Model unavailable — engine-written commentary used")
        sheet.cell(row, 1, state).font = Font(bold=True, size=10, color=NAVY)
        row += 2

        row = _header(sheet, row, ["Ref", "Heading", "Commentary", "Risk",
                                   "Source", "Numbers"])
        for s in self.narrative["sections"]:
            risk, check = s["risk_level"], s["number_check"]
            values = [s["section_code"], s["heading"], s["commentary"], risk,
                     _source_label(s["source"]), check]
            for i, v in enumerate(values, 1):
                cell = sheet.cell(row, i, v)
                cell.border = BORDER
                cell.font = Font(size=9)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if i == 4:
                    cell.fill = PatternFill("solid", fgColor=RISK_FILL.get(risk, GREY))
                    cell.font = Font(size=9, bold=True)
                    cell.alignment = Alignment(horizontal="center")
                if i == 6:
                    ok = check == "PASS"
                    cell.fill = PatternFill("solid", fgColor=STATUS_FILL["PASS" if ok else "FAIL"])
                    cell.font = Font(size=9, bold=True, color=STATUS_TEXT["PASS" if ok else "FAIL"])
                    cell.alignment = Alignment(horizontal="center")
            sheet.row_dimensions[row].height = 56
            row += 1

    def build(self) -> Workbook:
        wb = Workbook()
        self._cover(wb)
        self._checks(wb)
        self._narrative_sheet(wb)
        return wb

    def filename(self) -> str:
        company = "_".join("".join(c for c in self.review["company_name"]
                                   if c.isalnum() or c == " ").split())[:40]
        return f"WP514_{company}.xlsx"

    def save(self, path: str) -> str:
        if os.path.isdir(path) or path.endswith(os.sep):
            os.makedirs(path, exist_ok=True)
            path = os.path.join(path, self.filename())
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        self.build().save(path)
        return path


def generate_wp514(filepath, output_path=None):
    from backend.ingestion.loader import load_statements
    from backend.review.engine import StatementReviewEngine
    from backend.review.narrative import StatementNarrativeAgent
    data = load_statements(filepath)
    review = StatementReviewEngine(data).run()
    narrative = StatementNarrativeAgent().run(review)
    gen = StatementWP514Generator(review, narrative)
    return gen.save(output_path or gen.filename())


if __name__ == "__main__":
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument("filepath")
    p.add_argument("--out", default="out/")
    args = p.parse_args()
    print("Wrote", generate_wp514(args.filepath, args.out))